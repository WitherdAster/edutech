from datetime import date, datetime
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.auth import (
    get_db,
    get_current_user,
    hash_password,
    verify_password,
    create_access_token,
)
from app.models import User, Guru, Student, Kelas, Jurusan, Attendance, GuruKelas, MataPelajaran, Jadwal

router = APIRouter(prefix="/api", tags=["API"])


class LoginRequest(BaseModel):
    username: str
    password: str


class AbsensiUpdate(BaseModel):
    status_manual: str
    keterangan: str | None = None


class CreateGuruRequest(BaseModel):
    username: str
    password: str
    nama: str
    nip: str | None = None
    jenis_kelamin: str | None = None
    tempat_lahir: str | None = None
    tanggal_lahir: str | None = None
    agama: str | None = None
    alamat: str | None = None
    no_telp: str | None = None


class UpdateGuruRequest(BaseModel):
    nama: str | None = None
    nip: str | None = None
    jenis_kelamin: str | None = None
    tempat_lahir: str | None = None
    tanggal_lahir: str | None = None
    agama: str | None = None
    alamat: str | None = None
    no_telp: str | None = None


class CreateSiswaRequest(BaseModel):
    nisn: str
    nama_siswa: str
    id_kelas: int
    jenis_kelamin: str | None = None
    tempat_lahir: str | None = None
    tanggal_lahir: str | None = None
    agama: str | None = None
    alamat: str | None = None
    no_telp: str | None = None
    tahun_masuk: int | None = None


class UpdateSiswaRequest(BaseModel):
    nisn: str | None = None
    nama_siswa: str | None = None
    id_kelas: int | None = None
    jenis_kelamin: str | None = None
    tempat_lahir: str | None = None
    tanggal_lahir: str | None = None
    agama: str | None = None
    alamat: str | None = None
    no_telp: str | None = None
    tahun_masuk: int | None = None


class CreateMapelRequest(BaseModel):
    nama_mapel: str


class CreateJurusanRequest(BaseModel):
    nama_jurusan: str


class CreateKelasRequest(BaseModel):
    nama_kelas: str
    id_jurusan: int


class CreateJadwalRequest(BaseModel):
    id_mapel: int
    id_user: int
    id_kelas: int
    hari: str
    jam_mulai: str
    jam_selesai: str


class UpdateJadwalRequest(BaseModel):
    id_mapel: int | None = None
    id_user: int | None = None
    jam_mulai: str | None = None
    jam_selesai: str | None = None


@router.post("/auth/login")
def login(req: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == req.username).first()
    if not user or not verify_password(req.password, user.password):
        raise HTTPException(status_code=401, detail="Username atau password salah")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Akun tidak aktif")

    token = create_access_token({"sub": user.id_user})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id_user": user.id_user,
            "username": user.username,
            "nama": user.nama,
            "role": user.role,
        },
    }


@router.post("/utils/fix-passwords")
def fix_passwords(db: Session = Depends(get_db)):
    count = 0
    for u in db.query(User).all():
        if not u.password.startswith("$2"):
            u.password = hash_password(u.password)
            count += 1
            print(f"Fixed: {u.username}")
    db.commit()
    return {"fixed": count}


@router.get("/me")
def get_me(user: User = Depends(get_current_user)):
    return {
        "id_user": user.id_user,
        "username": user.username,
        "nama": user.nama,
        "role": user.role,
    }


@router.get("/kelas")
def list_kelas(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = (
        db.query(Kelas)
        .options(joinedload(Kelas.jurusan))
        .order_by(Kelas.nama_kelas)
    )

    if user.role == "guru":
        subquery = db.query(GuruKelas.id_kelas).filter(GuruKelas.id_user == user.id_user).subquery()
        query = query.filter(Kelas.id_kelas.in_(subquery))

    kelas_list = query.all()
    return [
        {
            "id_kelas": k.id_kelas,
            "nama_kelas": k.nama_kelas,
            "id_jurusan": k.id_jurusan,
            "jurusan": k.jurusan.nama_jurusan if k.jurusan else None,
        }
        for k in kelas_list
    ]


@router.get("/jurusan")
def list_jurusan(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    jurusan_list = db.query(Jurusan).order_by(Jurusan.nama_jurusan).all()
    return [
        {"id_jurusan": j.id_jurusan, "nama_jurusan": j.nama_jurusan}
        for j in jurusan_list
    ]


@router.post("/admin/jurusan")
def create_jurusan(
    body: CreateJurusanRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat menambah data jurusan")

    jurusan = Jurusan(nama_jurusan=body.nama_jurusan)
    db.add(jurusan)
    db.commit()
    db.refresh(jurusan)
    return {"id_jurusan": jurusan.id_jurusan, "nama_jurusan": jurusan.nama_jurusan}


@router.put("/admin/jurusan/{id_jurusan}")
def update_jurusan(
    id_jurusan: int,
    body: CreateJurusanRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat mengubah data jurusan")

    jurusan = db.query(Jurusan).filter(Jurusan.id_jurusan == id_jurusan).first()
    if not jurusan:
        raise HTTPException(status_code=404, detail="Jurusan tidak ditemukan")

    jurusan.nama_jurusan = body.nama_jurusan
    db.commit()
    return {"message": "Jurusan berhasil diperbarui"}


@router.delete("/admin/jurusan/{id_jurusan}")
def delete_jurusan(
    id_jurusan: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat menghapus data jurusan")

    jurusan = db.query(Jurusan).filter(Jurusan.id_jurusan == id_jurusan).first()
    if not jurusan:
        raise HTTPException(status_code=404, detail="Jurusan tidak ditemukan")

    db.delete(jurusan)
    db.commit()
    return {"message": "Jurusan berhasil dihapus"}


@router.get("/siswa")
def list_siswa(
    id_kelas: int | None = Query(None),
    search: str | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = db.query(Student).options(joinedload(Student.kelas_rel).joinedload(Kelas.jurusan))

    if user.role == "guru":
        allowed_kelas = db.query(GuruKelas.id_kelas).filter(GuruKelas.id_user == user.id_user).subquery()
        query = query.filter(Student.id_kelas.in_(allowed_kelas))

    if id_kelas:
        query = query.filter(Student.id_kelas == id_kelas)
    if search:
        query = query.filter(Student.nama_siswa.ilike(f"%{search}%"))

    siswa_list = query.order_by(Student.nama_siswa).all()
    return [
        {
            "id_siswa": s.id_siswa,
            "nisn": s.nisn,
            "nama_siswa": s.nama_siswa,
            "jenis_kelamin": s.jenis_kelamin,
            "tempat_lahir": s.tempat_lahir,
            "tanggal_lahir": s.tanggal_lahir.isoformat() if s.tanggal_lahir else None,
            "agama": s.agama,
            "alamat": s.alamat,
            "no_telp": s.no_telp,
            "tahun_masuk": s.tahun_masuk,
            "kelas": s.kelas_rel.nama_kelas if s.kelas_rel else None,
            "jurusan": s.kelas_rel.jurusan.nama_jurusan if s.kelas_rel and s.kelas_rel.jurusan else None,
        }
        for s in siswa_list
    ]


@router.get("/siswa/{id_siswa}")
def detail_siswa(
    id_siswa: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    student = (
        db.query(Student)
        .options(joinedload(Student.kelas_rel).joinedload(Kelas.jurusan))
        .filter(Student.id_siswa == id_siswa)
        .first()
    )
    if not student:
        raise HTTPException(status_code=404, detail="Siswa tidak ditemukan")

    if user.role == "guru":
        allowed = (
            db.query(GuruKelas)
            .filter(
                GuruKelas.id_user == user.id_user,
                GuruKelas.id_kelas == student.id_kelas,
            )
            .first()
        )
        if not allowed:
            raise HTTPException(status_code=403, detail="Akses ditolak")

    return {
        "id_siswa": student.id_siswa,
        "nisn": student.nisn,
        "nama_siswa": student.nama_siswa,
        "id_kelas": student.id_kelas,
        "jenis_kelamin": student.jenis_kelamin,
        "tempat_lahir": student.tempat_lahir,
        "tanggal_lahir": student.tanggal_lahir.isoformat() if student.tanggal_lahir else None,
        "agama": student.agama,
        "alamat": student.alamat,
        "no_telp": student.no_telp,
        "tahun_masuk": student.tahun_masuk,
        "kelas": student.kelas_rel.nama_kelas if student.kelas_rel else None,
        "jurusan": student.kelas_rel.jurusan.nama_jurusan if student.kelas_rel and student.kelas_rel.jurusan else None,
    }


@router.get("/absensi")
def list_absensi(
    tanggal: str | None = Query(None),
    id_kelas: int | None = Query(None),
    id_siswa: int | None = Query(None),
    id_jadwal: int | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if tanggal:
        try:
            parsed = datetime.strptime(tanggal, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Format tanggal salah (YYYY-MM-DD)")
    else:
        parsed = date.today()

    day_start = datetime(parsed.year, parsed.month, parsed.day)
    day_end = datetime(parsed.year, parsed.month, parsed.day + 1)

    query = (
        db.query(Attendance)
        .options(
            joinedload(Attendance.siswa_rel).joinedload(Student.kelas_rel).joinedload(Kelas.jurusan),
            joinedload(Attendance.jadwal_rel).joinedload(Jadwal.mapel_rel),
        )
        .filter(
            Attendance.check_time >= day_start,
            Attendance.check_time < day_end,
        )
    )

    if user.role == "guru":
        allowed_kelas = db.query(GuruKelas.id_kelas).filter(GuruKelas.id_user == user.id_user).subquery()
        allowed_siswa = db.query(Student.id_siswa).filter(Student.id_kelas.in_(allowed_kelas)).subquery()
        query = query.filter(Attendance.id_siswa.in_(allowed_siswa))

    if id_kelas:
        filtered_siswa = db.query(Student.id_siswa).filter(Student.id_kelas == id_kelas).subquery()
        query = query.filter(Attendance.id_siswa.in_(filtered_siswa))

    if id_siswa:
        query = query.filter(Attendance.id_siswa == id_siswa)

    if id_jadwal:
        query = query.filter(Attendance.id_jadwal == id_jadwal)
    else:
        query = query.filter(Attendance.id_jadwal == None)

    absensi_list = query.order_by(Attendance.check_time.desc()).all()
    result = []
    for a in absensi_list:
        status_display = a.status_manual if a.status_manual else a.status
        row = {
            "id_absensi": a.id_absensi,
            "id_siswa": a.id_siswa,
            "nama_siswa": a.siswa_rel.nama_siswa if a.siswa_rel else "Unknown",
            "id_kelas": a.siswa_rel.id_kelas if a.siswa_rel else None,
            "kelas": a.siswa_rel.kelas_rel.nama_kelas if a.siswa_rel and a.siswa_rel.kelas_rel else None,
            "check_time": a.check_time.isoformat() if a.check_time else None,
            "status": status_display,
            "status_manual": a.status_manual,
            "keterangan": a.keterangan,
            "similarity": a.similarity,
            "updated_by": a.updated_by,
            "updated_by_name": a.updated_by_rel.nama if a.updated_by_rel else None,
            "id_jadwal": a.id_jadwal,
            "mata_pelajaran": a.jadwal_rel.mapel_rel.nama_mapel if a.jadwal_rel and a.jadwal_rel.mapel_rel else None,
            "sumber": "Base" if a.id_jadwal is None else "Manual",
        }
        result.append(row)
    return result


@router.post("/absensi")
def create_absensi(
    body: AbsensiUpdate,
    id_siswa: int = Query(...),
    id_jadwal: int = Query(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    allowed_status = {"Hadir", "Izin", "Sakit", "Alpa"}
    if body.status_manual not in allowed_status:
        raise HTTPException(status_code=400, detail=f"Status harus salah satu dari: {', '.join(sorted(allowed_status))}")

    siswa = db.query(Student).filter(Student.id_siswa == id_siswa).first()
    if not siswa:
        raise HTTPException(status_code=404, detail="Siswa tidak ditemukan")

    if user.role == "guru":
        allowed = (
            db.query(GuruKelas)
            .filter(
                GuruKelas.id_user == user.id_user,
                GuruKelas.id_kelas == siswa.id_kelas,
            )
            .first()
        )
        if not allowed:
            raise HTTPException(status_code=403, detail="Anda tidak memiliki akses ke siswa ini")

    existing = db.query(Attendance).filter(
        Attendance.id_siswa == id_siswa,
        Attendance.id_jadwal == id_jadwal,
    ).first()

    if existing:
        existing.status_manual = body.status_manual
        existing.keterangan = body.keterangan
        existing.updated_by = user.id_user
        existing.updated_at = datetime.now()
        db.commit()
        return {"message": "Status berhasil diperbarui", "id_absensi": existing.id_absensi}

    attendance = Attendance(
        id_siswa=id_siswa,
        id_jadwal=id_jadwal,
        check_time=datetime.now(),
        status_manual=body.status_manual,
        status=body.status_manual,
        keterangan=body.keterangan,
        updated_by=user.id_user,
        updated_at=datetime.now(),
    )
    db.add(attendance)
    db.commit()
    db.refresh(attendance)
    return {"message": "Status berhasil dibuat", "id_absensi": attendance.id_absensi}


@router.put("/absensi/{id_absensi}")
def update_absensi(
    id_absensi: int,
    body: AbsensiUpdate,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    allowed_status = {"Hadir", "Izin", "Sakit", "Alpa"}
    if body.status_manual not in allowed_status:
        raise HTTPException(status_code=400, detail=f"Status harus salah satu dari: {', '.join(sorted(allowed_status))}")

    attendance = db.query(Attendance).options(joinedload(Attendance.siswa_rel)).filter(
        Attendance.id_absensi == id_absensi
    ).first()
    if not attendance:
        raise HTTPException(status_code=404, detail="Data absensi tidak ditemukan")

    if user.role == "guru":
        allowed = (
            db.query(GuruKelas)
            .filter(
                GuruKelas.id_user == user.id_user,
                GuruKelas.id_kelas == attendance.siswa_rel.id_kelas,
            )
            .first()
        )
        if not allowed:
            raise HTTPException(status_code=403, detail="Anda tidak memiliki akses ke siswa ini")

    attendance.status_manual = body.status_manual
    attendance.keterangan = body.keterangan
    attendance.updated_by = user.id_user
    attendance.updated_at = datetime.now()
    db.commit()

    return {"message": "Status berhasil diperbarui", "id_absensi": id_absensi}


@router.get("/absensi/export")
def export_absensi(
    tanggal: str | None = Query(None),
    id_kelas: int | None = Query(None),
    id_mapel: int | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat mengexport data")

    if tanggal:
        try:
            parsed = datetime.strptime(tanggal, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Format tanggal salah (YYYY-MM-DD)")
    else:
        parsed = date.today()

    day_start = datetime(parsed.year, parsed.month, parsed.day)
    day_end = datetime(parsed.year, parsed.month, parsed.day + 1)

    query_records = (
        db.query(Attendance)
        .options(
            joinedload(Attendance.siswa_rel).joinedload(Student.kelas_rel).joinedload(Kelas.jurusan),
            joinedload(Attendance.jadwal_rel).joinedload(Jadwal.mapel_rel),
            joinedload(Attendance.jadwal_rel).joinedload(Jadwal.user_rel),
        )
        .filter(
            Attendance.check_time >= day_start,
            Attendance.check_time < day_end,
        )
    )

    if id_kelas:
        query_records = query_records.join(Student).filter(Student.id_kelas == id_kelas)

    if id_mapel:
        jadwal_ids = db.query(Jadwal.id_jadwal).filter(Jadwal.id_mapel == id_mapel).subquery()
        query_records = query_records.filter(Attendance.id_jadwal.in_(jadwal_ids))

    records = query_records.order_by(Attendance.check_time.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Absensi"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["No", "NISN", "Nama Siswa", "Kelas", "Jurusan", "Mata Pelajaran", "Guru Pengajar", "Waktu", "Status", "Keterangan"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for idx, rec in enumerate(records, 1):
        status_display = rec.status_manual if rec.status_manual else rec.status
        siswa = rec.siswa_rel
        kelas_rel = siswa.kelas_rel if siswa else None
        mapel_name = rec.jadwal_rel.mapel_rel.nama_mapel if rec.jadwal_rel and rec.jadwal_rel.mapel_rel else ("-" if rec.id_jadwal else "Base")
        guru_name = rec.jadwal_rel.user_rel.nama if rec.jadwal_rel and rec.jadwal_rel.user_rel else "-"
        row = [
            idx,
            siswa.nisn if siswa else "",
            siswa.nama_siswa if siswa else "",
            kelas_rel.nama_kelas if kelas_rel else "",
            kelas_rel.jurusan.nama_jurusan if kelas_rel and kelas_rel.jurusan else "",
            mapel_name,
            guru_name,
            rec.check_time.strftime("%Y-%m-%d %H:%M:%S") if rec.check_time else "",
            status_display,
            rec.keterangan or "",
        ]
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=idx + 1, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    col_letters = ["A","B","C","D","E","F","G","H","I","J"]
    for i, letter in enumerate(col_letters[:len(headers)]):
        ws.column_dimensions[letter].width = 20
    ws.column_dimensions["C"].width = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"rekap_absensi_{tanggal or date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/guru")
def list_guru(
    id_kelas: int | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = (
        db.query(Guru)
        .options(
            joinedload(Guru.user_rel),
            joinedload(Guru.user_rel).joinedload(User.guru_kelas_list).joinedload(GuruKelas.kelas_rel).joinedload(Kelas.jurusan),
        )
    )

    if id_kelas is not None:
        allowed_user_ids = (
            db.query(GuruKelas.id_user)
            .filter(GuruKelas.id_kelas == id_kelas)
            .subquery()
        )
        query = query.filter(Guru.id_user.in_(allowed_user_ids))

    guru_list = query.all()
    return [
        {
            "id_guru": g.id_guru,
            "id_user": g.id_user,
            "nip": g.nip,
            "nama": g.user_rel.nama if g.user_rel else None,
            "username": g.user_rel.username if g.user_rel else None,
            "jenis_kelamin": g.jenis_kelamin,
            "tempat_lahir": g.tempat_lahir,
            "tanggal_lahir": g.tanggal_lahir.isoformat() if g.tanggal_lahir else None,
            "agama": g.agama,
            "alamat": g.alamat,
            "no_telp": g.no_telp,
            "kelas_list": [
                {
                    "id_kelas": gk.id_kelas,
                    "nama_kelas": gk.kelas_rel.nama_kelas if gk.kelas_rel else None,
                    "jurusan": gk.kelas_rel.jurusan.nama_jurusan if gk.kelas_rel and gk.kelas_rel.jurusan else None,
                }
                for gk in (g.user_rel.guru_kelas_list if g.user_rel else [])
            ] if g.user_rel else [],
        }
        for g in guru_list
    ]


@router.get("/guru/{id_guru}")
def detail_guru(
    id_guru: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    guru = (
        db.query(Guru)
        .options(
            joinedload(Guru.user_rel),
            joinedload(Guru.user_rel).joinedload(User.guru_kelas_list).joinedload(GuruKelas.kelas_rel).joinedload(Kelas.jurusan),
        )
        .filter(Guru.id_guru == id_guru)
        .first()
    )
    if not guru:
        raise HTTPException(status_code=404, detail="Guru tidak ditemukan")
    return {
        "id_guru": guru.id_guru,
        "id_user": guru.id_user,
        "nip": guru.nip,
        "nama": guru.user_rel.nama if guru.user_rel else None,
        "username": guru.user_rel.username if guru.user_rel else None,
        "jenis_kelamin": guru.jenis_kelamin,
        "tempat_lahir": guru.tempat_lahir,
        "tanggal_lahir": guru.tanggal_lahir.isoformat() if guru.tanggal_lahir else None,
        "agama": guru.agama,
        "alamat": guru.alamat,
        "no_telp": guru.no_telp,
        "kelas_list": [
            {
                "id_kelas": gk.id_kelas,
                "nama_kelas": gk.kelas_rel.nama_kelas if gk.kelas_rel else None,
                "jurusan": gk.kelas_rel.jurusan.nama_jurusan if gk.kelas_rel and gk.kelas_rel.jurusan else None,
            }
            for gk in (guru.user_rel.guru_kelas_list if guru.user_rel else [])
        ] if guru.user_rel else [],
    }


@router.post("/admin/guru")
def create_guru(
    body: CreateGuruRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat menambah data guru")

    if db.query(User).filter(User.username == body.username).first():
        raise HTTPException(status_code=400, detail="Username sudah digunakan")

    tanggal = None
    if body.tanggal_lahir:
        try:
            tanggal = datetime.strptime(body.tanggal_lahir, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Format tanggal_lahir harus YYYY-MM-DD")

    new_user = User(
        username=body.username,
        password=hash_password(body.password),
        role="guru",
        nama=body.nama,
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    new_guru = Guru(
        id_user=new_user.id_user,
        nip=body.nip,
        jenis_kelamin=body.jenis_kelamin,
        tempat_lahir=body.tempat_lahir,
        tanggal_lahir=tanggal,
        agama=body.agama,
        alamat=body.alamat,
        no_telp=body.no_telp,
    )
    db.add(new_guru)
    db.commit()
    db.refresh(new_guru)

    return {"message": "Guru berhasil ditambahkan", "id_guru": new_guru.id_guru}


@router.put("/admin/guru/{id_guru}")
def update_guru(
    id_guru: int,
    body: UpdateGuruRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat mengubah data guru")

    guru = db.query(Guru).options(joinedload(Guru.user_rel)).filter(Guru.id_guru == id_guru).first()
    if not guru:
        raise HTTPException(status_code=404, detail="Guru tidak ditemukan")

    if body.nama is not None:
        guru.user_rel.nama = body.nama
    if body.nip is not None:
        guru.nip = body.nip
    if body.jenis_kelamin is not None:
        guru.jenis_kelamin = body.jenis_kelamin
    if body.tempat_lahir is not None:
        guru.tempat_lahir = body.tempat_lahir
    if body.tanggal_lahir is not None:
        try:
            guru.tanggal_lahir = datetime.strptime(body.tanggal_lahir, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Format tanggal_lahir harus YYYY-MM-DD")
    if body.agama is not None:
        guru.agama = body.agama
    if body.alamat is not None:
        guru.alamat = body.alamat
    if body.no_telp is not None:
        guru.no_telp = body.no_telp

    db.commit()
    return {"message": "Guru berhasil diperbarui"}


@router.delete("/admin/guru/{id_guru}")
def delete_guru(
    id_guru: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat menghapus data guru")

    guru = db.query(Guru).options(joinedload(Guru.user_rel)).filter(Guru.id_guru == id_guru).first()
    if not guru:
        raise HTTPException(status_code=404, detail="Guru tidak ditemukan")

    guru.user_rel.is_active = False
    db.commit()
    return {"message": "Guru berhasil dinonaktifkan"}


@router.get("/admin/guru/{id_guru}/kelas")
def list_guru_kelas(
    id_guru: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat mengelola kelas guru")

    guru = db.query(Guru).filter(Guru.id_guru == id_guru).first()
    if not guru:
        raise HTTPException(status_code=404, detail="Guru tidak ditemukan")

    assignments = (
        db.query(GuruKelas)
        .options(joinedload(GuruKelas.kelas_rel).joinedload(Kelas.jurusan))
        .filter(GuruKelas.id_user == guru.id_user)
        .all()
    )
    return [
        {
            "id": a.id,
            "id_kelas": a.id_kelas,
            "nama_kelas": a.kelas_rel.nama_kelas if a.kelas_rel else None,
            "jurusan": a.kelas_rel.jurusan.nama_jurusan if a.kelas_rel and a.kelas_rel.jurusan else None,
        }
        for a in assignments
    ]


class AssignKelasRequest(BaseModel):
    id_kelas: int


@router.post("/admin/guru/{id_guru}/kelas")
def assign_guru_kelas(
    id_guru: int,
    body: AssignKelasRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat mengelola kelas guru")

    guru = db.query(Guru).filter(Guru.id_guru == id_guru).first()
    if not guru:
        raise HTTPException(status_code=404, detail="Guru tidak ditemukan")

    kelas = db.query(Kelas).filter(Kelas.id_kelas == body.id_kelas).first()
    if not kelas:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")

    existing = (
        db.query(GuruKelas)
        .filter(
            GuruKelas.id_user == guru.id_user,
            GuruKelas.id_kelas == body.id_kelas,
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="Kelas sudah ditambahkan untuk guru ini")

    assign = GuruKelas(id_user=guru.id_user, id_kelas=body.id_kelas)
    db.add(assign)
    db.commit()

    return {"message": "Kelas berhasil ditambahkan untuk guru"}


@router.delete("/admin/guru/{id_guru}/kelas/{id_kelas}")
def remove_guru_kelas(
    id_guru: int,
    id_kelas: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat mengelola kelas guru")

    guru = db.query(Guru).filter(Guru.id_guru == id_guru).first()
    if not guru:
        raise HTTPException(status_code=404, detail="Guru tidak ditemukan")

    assign = (
        db.query(GuruKelas)
        .filter(
            GuruKelas.id_user == guru.id_user,
            GuruKelas.id_kelas == id_kelas,
        )
        .first()
    )
    if not assign:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan untuk guru ini")

    db.delete(assign)
    db.commit()

    return {"message": "Kelas berhasil dihapus dari guru"}


@router.post("/admin/siswa")
def create_siswa(
    body: CreateSiswaRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat menambah data siswa")

    if db.query(Student).filter(Student.nisn == body.nisn).first():
        raise HTTPException(status_code=400, detail="NISN sudah digunakan")

    if not db.query(Kelas).filter(Kelas.id_kelas == body.id_kelas).first():
        raise HTTPException(status_code=400, detail="Kelas tidak ditemukan")

    tanggal = None
    if body.tanggal_lahir:
        try:
            tanggal = datetime.strptime(body.tanggal_lahir, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Format tanggal_lahir harus YYYY-MM-DD")

    siswa = Student(
        nisn=body.nisn,
        nama_siswa=body.nama_siswa,
        id_kelas=body.id_kelas,
        jenis_kelamin=body.jenis_kelamin,
        tempat_lahir=body.tempat_lahir,
        tanggal_lahir=tanggal,
        agama=body.agama,
        alamat=body.alamat,
        no_telp=body.no_telp,
        tahun_masuk=body.tahun_masuk,
    )
    db.add(siswa)
    db.commit()
    db.refresh(siswa)

    return {"message": "Siswa berhasil ditambahkan", "id_siswa": siswa.id_siswa}


@router.put("/admin/siswa/{id_siswa}")
def update_siswa(
    id_siswa: int,
    body: UpdateSiswaRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat mengubah data siswa")

    siswa = db.query(Student).filter(Student.id_siswa == id_siswa).first()
    if not siswa:
        raise HTTPException(status_code=404, detail="Siswa tidak ditemukan")

    if body.nisn is not None:
        existing = db.query(Student).filter(Student.nisn == body.nisn, Student.id_siswa != id_siswa).first()
        if existing:
            raise HTTPException(status_code=400, detail="NISN sudah digunakan")
        siswa.nisn = body.nisn
    if body.nama_siswa is not None:
        siswa.nama_siswa = body.nama_siswa
    if body.id_kelas is not None:
        if not db.query(Kelas).filter(Kelas.id_kelas == body.id_kelas).first():
            raise HTTPException(status_code=400, detail="Kelas tidak ditemukan")
        siswa.id_kelas = body.id_kelas
    if body.jenis_kelamin is not None:
        siswa.jenis_kelamin = body.jenis_kelamin
    if body.tempat_lahir is not None:
        siswa.tempat_lahir = body.tempat_lahir
    if body.tanggal_lahir is not None:
        try:
            siswa.tanggal_lahir = datetime.strptime(body.tanggal_lahir, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Format tanggal_lahir harus YYYY-MM-DD")
    if body.agama is not None:
        siswa.agama = body.agama
    if body.alamat is not None:
        siswa.alamat = body.alamat
    if body.no_telp is not None:
        siswa.no_telp = body.no_telp
    if body.tahun_masuk is not None:
        siswa.tahun_masuk = body.tahun_masuk

    db.commit()
    return {"message": "Siswa berhasil diperbarui"}


@router.delete("/admin/siswa/{id_siswa}")
def delete_siswa(
    id_siswa: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat menghapus data siswa")

    siswa = db.query(Student).filter(Student.id_siswa == id_siswa).first()
    if not siswa:
        raise HTTPException(status_code=404, detail="Siswa tidak ditemukan")

    db.delete(siswa)
    db.commit()
    return {"message": "Siswa berhasil dihapus"}


@router.get("/mapel")
def list_mapel(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    mapel_list = db.query(MataPelajaran).order_by(MataPelajaran.nama_mapel).all()
    return [
        {"id_mapel": m.id_mapel, "nama_mapel": m.nama_mapel}
        for m in mapel_list
    ]


@router.post("/admin/mapel")
def create_mapel(
    body: CreateMapelRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat menambah data mapel")

    mapel = MataPelajaran(nama_mapel=body.nama_mapel)
    db.add(mapel)
    db.commit()
    db.refresh(mapel)
    return {"id_mapel": mapel.id_mapel, "nama_mapel": mapel.nama_mapel}


@router.put("/admin/mapel/{id_mapel}")
def update_mapel(
    id_mapel: int,
    body: CreateMapelRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat mengubah data mapel")

    mapel = db.query(MataPelajaran).filter(MataPelajaran.id_mapel == id_mapel).first()
    if not mapel:
        raise HTTPException(status_code=404, detail="Mapel tidak ditemukan")

    mapel.nama_mapel = body.nama_mapel
    db.commit()
    return {"message": "Mapel berhasil diperbarui"}


@router.delete("/admin/mapel/{id_mapel}")
def delete_mapel(
    id_mapel: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat menghapus data mapel")

    mapel = db.query(MataPelajaran).filter(MataPelajaran.id_mapel == id_mapel).first()
    if not mapel:
        raise HTTPException(status_code=404, detail="Mapel tidak ditemukan")

    db.delete(mapel)
    db.commit()
    return {"message": "Mapel berhasil dihapus"}


@router.get("/jadwal")
def list_jadwal(
    hari: str | None = Query(None),
    id_kelas: int | None = Query(None),
    id_user: int | None = Query(None),
    id_jurusan: int | None = Query(None),
    id_mapel: int | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = (
        db.query(Jadwal)
        .options(
            joinedload(Jadwal.mapel_rel),
            joinedload(Jadwal.user_rel),
            joinedload(Jadwal.kelas_rel).joinedload(Kelas.jurusan),
        )
    )

    if user.role == "guru":
        query = query.filter(Jadwal.id_user == user.id_user)

    if hari:
        query = query.filter(Jadwal.hari == hari)
    if id_kelas:
        query = query.filter(Jadwal.id_kelas == id_kelas)
    if id_user:
        query = query.filter(Jadwal.id_user == id_user)
    if id_jurusan:
        query = query.join(Kelas).filter(Kelas.id_jurusan == id_jurusan)
    if id_mapel:
        query = query.filter(Jadwal.id_mapel == id_mapel)

    jadwal_list = query.order_by(Jadwal.hari, Jadwal.jam_mulai).all()
    return [
        {
            "id_jadwal": j.id_jadwal,
            "id_mapel": j.id_mapel,
            "nama_mapel": j.mapel_rel.nama_mapel if j.mapel_rel else None,
            "id_user": j.id_user,
            "nama_guru": j.user_rel.nama if j.user_rel else None,
            "id_kelas": j.id_kelas,
            "nama_kelas": j.kelas_rel.nama_kelas if j.kelas_rel else None,
            "jurusan": j.kelas_rel.jurusan.nama_jurusan if j.kelas_rel and j.kelas_rel.jurusan else None,
            "hari": j.hari,
            "jam_mulai": j.jam_mulai.strftime("%H:%M") if j.jam_mulai else None,
            "jam_selesai": j.jam_selesai.strftime("%H:%M") if j.jam_selesai else None,
        }
        for j in jadwal_list
    ]


@router.post("/admin/jadwal")
def create_jadwal(
    body: CreateJadwalRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat menambah data jadwal")

    if body.hari not in ("Senin", "Selasa", "Rabu", "Kamis", "Jumat"):
        raise HTTPException(status_code=400, detail="Hari tidak valid")

    try:
        jam_mulai = datetime.strptime(body.jam_mulai, "%H:%M").time()
        jam_selesai = datetime.strptime(body.jam_selesai, "%H:%M").time()
    except ValueError:
        raise HTTPException(status_code=400, detail="Format jam harus HH:MM")

    if jam_mulai >= jam_selesai:
        raise HTTPException(status_code=400, detail="jam_mulai harus sebelum jam_selesai")

    jadwal = Jadwal(
        id_mapel=body.id_mapel,
        id_user=body.id_user,
        id_kelas=body.id_kelas,
        hari=body.hari,
        jam_mulai=jam_mulai,
        jam_selesai=jam_selesai,
    )
    db.add(jadwal)
    db.commit()
    db.refresh(jadwal)
    return {"message": "Jadwal berhasil ditambahkan", "id_jadwal": jadwal.id_jadwal}


@router.put("/admin/jadwal/{id_jadwal}")
def update_jadwal(
    id_jadwal: int,
    body: UpdateJadwalRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat mengubah data jadwal")

    jadwal = db.query(Jadwal).filter(Jadwal.id_jadwal == id_jadwal).first()
    if not jadwal:
        raise HTTPException(status_code=404, detail="Jadwal tidak ditemukan")

    if body.id_mapel is not None:
        jadwal.id_mapel = body.id_mapel
    if body.id_user is not None:
        jadwal.id_user = body.id_user
    if body.jam_mulai is not None:
        try:
            jadwal.jam_mulai = datetime.strptime(body.jam_mulai, "%H:%M").time()
        except ValueError:
            raise HTTPException(status_code=400, detail="Format jam_mulai harus HH:MM")
    if body.jam_selesai is not None:
        try:
            jadwal.jam_selesai = datetime.strptime(body.jam_selesai, "%H:%M").time()
        except ValueError:
            raise HTTPException(status_code=400, detail="Format jam_selesai harus HH:MM")

    db.commit()
    return {"message": "Jadwal berhasil diperbarui"}


@router.delete("/admin/jadwal/{id_jadwal}")
def delete_jadwal(
    id_jadwal: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat menghapus data jadwal")

    jadwal = db.query(Jadwal).filter(Jadwal.id_jadwal == id_jadwal).first()
    if not jadwal:
        raise HTTPException(status_code=404, detail="Jadwal tidak ditemukan")

    db.delete(jadwal)
    db.commit()
    return {"message": "Jadwal berhasil dihapus"}


@router.get("/absensi/hari-ini")
def absensi_hari_ini(
    tanggal: str | None = Query(None),
    id_kelas: int | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if tanggal:
        try:
            today = datetime.strptime(tanggal, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Format tanggal salah (YYYY-MM-DD)")
    else:
        today = date.today()
    today_start = datetime(today.year, today.month, today.day)
    today_end = datetime(today.year, today.month, today.day + 1)
    hari_ini = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat"][today.weekday()]

    jadwal_query = (
        db.query(Jadwal)
        .options(
            joinedload(Jadwal.mapel_rel),
            joinedload(Jadwal.user_rel),
            joinedload(Jadwal.kelas_rel).joinedload(Kelas.jurusan),
        )
        .filter(Jadwal.hari == hari_ini)
    )

    if user.role == "guru":
        jadwal_query = jadwal_query.filter(Jadwal.id_user == user.id_user)
    if id_kelas:
        jadwal_query = jadwal_query.filter(Jadwal.id_kelas == id_kelas)

    jadwal_list = jadwal_query.order_by(Jadwal.jam_mulai).all()

    base_absensi = (
        db.query(Attendance)
        .options(joinedload(Attendance.siswa_rel).joinedload(Student.kelas_rel).joinedload(Kelas.jurusan))
        .filter(
            Attendance.check_time >= today_start,
            Attendance.check_time < today_end,
            Attendance.id_jadwal == None,
        )
        .all()
    )
    base_map = {a.id_siswa: a for a in base_absensi}

    result = []
    for j in jadwal_list:
        siswa_list = (
            db.query(Student)
            .options(joinedload(Student.kelas_rel).joinedload(Kelas.jurusan))
            .filter(Student.id_kelas == j.id_kelas)
            .order_by(Student.nama_siswa)
            .all()
        )

        jadwal_records = (
            db.query(Attendance)
            .options(joinedload(Attendance.jadwal_rel))
            .filter(
                Attendance.check_time >= today_start,
                Attendance.check_time < today_end,
                Attendance.id_jadwal == j.id_jadwal,
            )
            .all()
        )
        jadwal_map = {a.id_siswa: a for a in jadwal_records}

        siswa_data = []
        hadir_count = 0
        belum_count = 0

        for s in siswa_list:
            per_jadwal = jadwal_map.get(s.id_siswa)
            base = base_map.get(s.id_siswa)

            if per_jadwal:
                status_display = per_jadwal.status_manual if per_jadwal.status_manual else per_jadwal.status
                check_time = per_jadwal.check_time.isoformat() if per_jadwal.check_time else None
                sumber = "Manual"
                id_absensi = per_jadwal.id_absensi
            elif base:
                status_display = base.status_manual if base.status_manual else base.status
                check_time = base.check_time.isoformat() if base.check_time else None
                sumber = "Base"
                id_absensi = base.id_absensi
            else:
                status_display = "Belum Absen"
                check_time = None
                sumber = None
                id_absensi = None

            if status_display == "Hadir":
                hadir_count += 1
            elif status_display == "Belum Absen":
                belum_count += 1

            siswa_data.append({
                "id_siswa": s.id_siswa,
                "nisn": s.nisn,
                "nama_siswa": s.nama_siswa,
                "kelas": s.kelas_rel.nama_kelas if s.kelas_rel else None,
                "status": status_display,
                "check_time": check_time,
                "sumber": sumber,
                "id_absensi": id_absensi,
            })

        total_siswa = len(siswa_data)
        result.append({
            "id_jadwal": j.id_jadwal,
            "nama_mapel": j.mapel_rel.nama_mapel if j.mapel_rel else None,
            "nama_guru": j.user_rel.nama if j.user_rel else None,
            "id_kelas": j.id_kelas,
            "nama_kelas": j.kelas_rel.nama_kelas if j.kelas_rel else None,
            "hari": j.hari,
            "jam_mulai": j.jam_mulai.strftime("%H:%M") if j.jam_mulai else None,
            "jam_selesai": j.jam_selesai.strftime("%H:%M") if j.jam_selesai else None,
            "total_siswa": total_siswa,
            "hadir": hadir_count,
            "belum": belum_count,
            "siswa": siswa_data,
        })

    return result


@router.get("/dashboard")
def dashboard(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    today = date.today()
    today_start = datetime(today.year, today.month, today.day)
    today_end = datetime(today.year, today.month, today.day + 1)
    hari_ini = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat"][today.weekday()]

    if user.role == "guru":
        guru_kelas_ids = [
            gk.id_kelas
            for gk in db.query(GuruKelas).filter(GuruKelas.id_user == user.id_user).all()
        ]
        total_siswa = (
            db.query(Student)
            .filter(Student.id_kelas.in_(guru_kelas_ids))
            .count()
        )
        kelas_count = len(guru_kelas_ids)

        hadir_today = (
            db.query(Attendance)
            .join(Student)
            .filter(
                Student.id_kelas.in_(guru_kelas_ids),
                Attendance.check_time >= today_start,
                Attendance.check_time < today_end,
                Attendance.id_jadwal == None,
                Attendance.status == "Hadir",
            )
            .count()
        )
    else:
        total_siswa = db.query(Student).count()
        kelas_count = db.query(Kelas).count()
        hadir_today = (
            db.query(Attendance)
            .filter(
                Attendance.check_time >= today_start,
                Attendance.check_time < today_end,
                Attendance.id_jadwal == None,
                Attendance.status == "Hadir",
            )
            .count()
        )

    jadwal_query = (
        db.query(Jadwal)
        .options(
            joinedload(Jadwal.mapel_rel),
            joinedload(Jadwal.user_rel),
            joinedload(Jadwal.kelas_rel).joinedload(Kelas.jurusan),
        )
        .filter(Jadwal.hari == hari_ini)
    )
    if user.role == "guru":
        jadwal_query = jadwal_query.filter(Jadwal.id_user == user.id_user)

    jadwal_list = jadwal_query.order_by(Jadwal.jam_mulai).all()

    base_absensi = (
        db.query(Attendance)
        .filter(
            Attendance.check_time >= today_start,
            Attendance.check_time < today_end,
            Attendance.id_jadwal == None,
        )
        .all()
    )
    base_siswa_ids = set(a.id_siswa for a in base_absensi)
    belum_hadir = total_siswa - len(base_siswa_ids)

    jadwal_stats = []
    for j in jadwal_list:
        kelas_siswa = db.query(Student).filter(Student.id_kelas == j.id_kelas).count()

        per_jadwal_ids = set(
            a.id_siswa for a in db.query(Attendance).filter(
                Attendance.check_time >= today_start,
                Attendance.check_time < today_end,
                Attendance.id_jadwal == j.id_jadwal,
                Attendance.status_manual.in_(["Hadir", "Izin", "Sakit", "Alpa"]),
            ).all()
        )

        override_sakits = set(
            a.id_siswa for a in db.query(Attendance).filter(
                Attendance.check_time >= today_start,
                Attendance.check_time < today_end,
                Attendance.id_jadwal == j.id_jadwal,
                Attendance.status_manual.in_(["Sakit", "Izin", "Alpa"]),
            ).all()
        )

        hadir_jadwal = len(
            s for s in (base_siswa_ids | per_jadwal_ids)
            if s not in override_sakits
        )

        jadwal_stats.append({
            "id_jadwal": j.id_jadwal,
            "nama_mapel": j.mapel_rel.nama_mapel if j.mapel_rel else None,
            "nama_guru": j.user_rel.nama if j.user_rel else None,
            "jam_mulai": j.jam_mulai.strftime("%H:%M") if j.jam_mulai else None,
            "jam_selesai": j.jam_selesai.strftime("%H:%M") if j.jam_selesai else None,
            "total_siswa": kelas_siswa,
            "hadir": hadir_jadwal,
        })

    return {
        "total_siswa": total_siswa,
        "hadir_hari_ini": hadir_today,
        "belum_hadir": belum_hadir,
        "total_kelas": kelas_count,
        "jadwal_stats": jadwal_stats,
    }


@router.post("/admin/kelas")
def create_kelas(
    body: CreateKelasRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat menambah data kelas")

    kelas = Kelas(nama_kelas=body.nama_kelas, id_jurusan=body.id_jurusan)
    db.add(kelas)
    db.commit()
    db.refresh(kelas)
    return {"id_kelas": kelas.id_kelas, "nama_kelas": kelas.nama_kelas, "id_jurusan": kelas.id_jurusan}


@router.put("/admin/kelas/{id_kelas}")
def update_kelas(
    id_kelas: int,
    body: CreateKelasRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat mengubah data kelas")

    kelas = db.query(Kelas).filter(Kelas.id_kelas == id_kelas).first()
    if not kelas:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")

    kelas.nama_kelas = body.nama_kelas
    kelas.id_jurusan = body.id_jurusan
    db.commit()
    return {"message": "Kelas berhasil diperbarui"}


@router.delete("/admin/kelas/{id_kelas}")
def delete_kelas(
    id_kelas: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "tu":
        raise HTTPException(status_code=403, detail="Hanya TU yang dapat menghapus data kelas")

    kelas = db.query(Kelas).filter(Kelas.id_kelas == id_kelas).first()
    if not kelas:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")

    db.delete(kelas)
    db.commit()
    return {"message": "Kelas berhasil dihapus"}
